import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

# Setup workbook
wb = openpyxl.Workbook()

# Remove default sheet
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# Styles Configuration
font_name = "Segoe UI"
font_title = Font(name=font_name, size=16, bold=True, color="FFFFFF")
font_subtitle = Font(name=font_name, size=10, italic=True, color="E5E7E9")
font_section = Font(name=font_name, size=12, bold=True, color="1F4E78")
font_header = Font(name=font_name, size=11, bold=True, color="FFFFFF")
font_bold = Font(name=font_name, size=11, bold=True)
font_regular = Font(name=font_name, size=11)
font_kpi_val = Font(name=font_name, size=18, bold=True, color="1F4E78")
font_kpi_lbl = Font(name=font_name, size=9, bold=True, color="595959")

fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")
fill_kpi = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
fill_section = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")

thin_side = Side(style='thin', color='D9D9D9')
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_double_bottom = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side(style='double', color='1F4E78'))
border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side(style='medium', color='1F4E78'))

# Status Color Fills
status_styles = {
    "Completed": {
        "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "font": Font(name=font_name, size=11, color="375623", bold=True)
    },
    "In Progress": {
        "fill": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "font": Font(name=font_name, size=11, color="1F4E78", bold=True)
    },
    "Review": {
        "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "font": Font(name=font_name, size=11, color="7F6000", bold=True)
    },
    "Blocked": {
        "fill": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
        "font": Font(name=font_name, size=11, color="C65911", bold=True)
    },
    "Open": {
        "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        "font": Font(name=font_name, size=11, color="595959")
    }
}

priority_styles = {
    "High": {
        "fill": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
        "font": Font(name=font_name, size=11, color="78281F", bold=True)
    },
    "Medium": {
        "fill": PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
        "font": Font(name=font_name, size=11, color="7E5109", bold=True)
    },
    "Low": {
        "fill": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        "font": Font(name=font_name, size=11, color="1E8449")
    }
}

# ==========================================
# SHEET 1: DASHBOARD
# ==========================================
ws_dash = wb.create_sheet(title="Dashboard")
ws_dash.views.sheetView[0].showGridLines = True

# Title block
ws_dash.merge_cells("B2:G2")
ws_dash["B2"] = "NutriBite Project Tracking Dashboard"
ws_dash["B2"].font = font_title
ws_dash["B2"].fill = fill_header
ws_dash["B2"].alignment = Alignment(horizontal="center", vertical="center")

ws_dash.merge_cells("B3:G3")
ws_dash["B3"] = "MERN + FastAPI Pediatric Nutrition Platform - Status Registry"
ws_dash["B3"].font = font_subtitle
ws_dash["B3"].fill = fill_header
ws_dash["B3"].alignment = Alignment(horizontal="center", vertical="center")

ws_dash.row_dimensions[2].height = 30
ws_dash.row_dimensions[3].height = 20

# KPI Cards definition
kpis = [
    ("Total Features", "=COUNTA(Features!$B$4:$B$17)", "B"),
    ("Completed Features", '=COUNTIF(Features!$F$4:$F$17, "Completed")', "C"),
    ("In Progress Features", '=COUNTIF(Features!$F$4:$F$17, "In Progress")', "D"),
    ("Open Features", '=COUNTIF(Features!$F$4:$F$17, "Open")', "E"),
    ("Blocked Features", '=COUNTIF(Features!$F$4:$F$17, "Blocked")', "F"),
    ("Overall Completion %", "=AVERAGE(Features!$H$4:$H$17)", "G")
]

for label, formula, col in kpis:
    cell_lbl = ws_dash[f"{col}5"]
    cell_lbl.value = label
    cell_lbl.font = font_kpi_lbl
    cell_lbl.fill = fill_kpi
    cell_lbl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_lbl.border = border_all
    
    cell_val = ws_dash[f"{col}6"]
    cell_val.value = formula
    cell_val.font = font_kpi_val
    cell_val.fill = fill_kpi
    cell_val.alignment = Alignment(horizontal="center", vertical="center")
    cell_val.border = border_all
    if label == "Overall Completion %":
        cell_val.number_format = "0.0%"

ws_dash.row_dimensions[5].height = 25
ws_dash.row_dimensions[6].height = 35

# Project Completion Progress Bar (Rows 8-9)
ws_dash.merge_cells("B8:G8")
ws_dash["B8"] = "Project Completion Progress Indicator"
ws_dash["B8"].font = Font(name=font_name, size=11, bold=True, color="1F4E78")
ws_dash["B8"].alignment = Alignment(horizontal="center", vertical="center")

ws_dash.merge_cells("B9:G9")
ws_dash["B9"] = "=G6"
ws_dash["B9"].font = Font(name=font_name, size=24, bold=True, color="1F4E78")
ws_dash["B9"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["B9"].number_format = "0.0%"
ws_dash["B9"].fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

for r in range(8, 10):
    for c in range(2, 8):
        ws_dash.cell(row=r, column=c).border = border_all

ws_dash.row_dimensions[8].height = 20
ws_dash.row_dimensions[9].height = 40

# Section Headers
ws_dash.merge_cells("B11:C11")
ws_dash["B11"] = "STATUS DISTRIBUTION DATA"
ws_dash["B11"].font = font_section
ws_dash["B11"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["B11"].fill = fill_section

ws_dash.merge_cells("E11:G11")
ws_dash["E11"] = "FEATURES STATUS DISTRIBUTION CHART"
ws_dash["E11"].font = font_section
ws_dash["E11"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["E11"].fill = fill_section

for c in [2, 3, 5, 6, 7]:
    ws_dash.cell(row=11, column=c).border = border_all

ws_dash.row_dimensions[11].height = 25

# Distribution Data Table
ws_dash["B12"] = "Status"
ws_dash["C12"] = "Count"
for col in ["B", "C"]:
    ws_dash[f"{col}12"].font = font_bold
    ws_dash[f"{col}12"].border = border_all
    ws_dash[f"{col}12"].alignment = Alignment(horizontal="center")

status_rows = [
    ("Completed", '=COUNTIF(Features!$F$4:$F$17, "Completed")'),
    ("In Progress", '=COUNTIF(Features!$F$4:$F$17, "In Progress")'),
    ("Review", '=COUNTIF(Features!$F$4:$F$17, "Review")'),
    ("Blocked", '=COUNTIF(Features!$F$4:$F$17, "Blocked")'),
    ("Open", '=COUNTIF(Features!$F$4:$F$17, "Open")')
]

for idx, (status_val, formula_val) in enumerate(status_rows, start=13):
    ws_dash[f"B{idx}"] = status_val
    ws_dash[f"B{idx}"].font = font_regular
    ws_dash[f"B{idx}"].border = border_all
    ws_dash[f"B{idx}"].alignment = Alignment(horizontal="left")
    
    ws_dash[f"C{idx}"] = formula_val
    ws_dash[f"C{idx}"].font = font_regular
    ws_dash[f"C{idx}"].border = border_all
    ws_dash[f"C{idx}"].alignment = Alignment(horizontal="center")

# Summary Totals
ws_dash["B18"] = "Total Features"
ws_dash["B18"].font = font_bold
ws_dash["B18"].border = border_double_bottom

ws_dash["C18"] = "=SUM(C13:C17)"
ws_dash["C18"].font = font_bold
ws_dash["C18"].border = border_double_bottom
ws_dash["C18"].alignment = Alignment(horizontal="center")

# Pie Chart setup
pie = PieChart()
labels = Reference(ws_dash, min_col=2, min_row=13, max_row=17)
data = Reference(ws_dash, min_col=3, min_row=12, max_row=17)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Features Status Distribution"
pie.width = 14
pie.height = 7
ws_dash.add_chart(pie, "E12")

# Project Health Summary Section
ws_dash.merge_cells("B20:G20")
ws_dash["B20"] = "PROJECT HEALTH & TIMELINE SUMMARY"
ws_dash["B20"].font = font_section
ws_dash["B20"].alignment = Alignment(horizontal="center", vertical="center")
ws_dash["B20"].fill = fill_section
for c in range(2, 8):
    ws_dash.cell(row=20, column=c).border = border_all

health_data = [
    ("Project Status:", "ACTIVE (Phase 5 - AI & Telemetry Integration)"),
    ("Current Health:", "HEALTHY (All core user workflows verified)"),
    ("Target Deadline:", "June 25, 2026 (Final Review)"),
    ("Risks/Blockers:", "None. Final RAG benchmark tests are complete.")
]

for idx, (label_h, val_h) in enumerate(health_data, start=21):
    ws_dash[f"B{idx}"] = label_h
    ws_dash[f"B{idx}"].font = font_bold
    ws_dash[f"B{idx}"].border = border_all
    ws_dash[f"B{idx}"].alignment = Alignment(horizontal="right")
    
    ws_dash.merge_cells(f"C{idx}:G{idx}")
    ws_dash[f"C{idx}"] = val_h
    ws_dash[f"C{idx}"].font = font_regular
    ws_dash[f"C{idx}"].alignment = Alignment(horizontal="left")
    for col_idx in range(3, 8):
        ws_dash.cell(row=idx, column=col_idx).border = border_all

# ==========================================
# SHEET 2: FEATURES
# ==========================================
ws_feat = wb.create_sheet(title="Features")
ws_feat.views.sheetView[0].showGridLines = True

# Title block
ws_feat.merge_cells("B2:I2")
ws_feat["B2"] = "NutriBite Feature Registry"
ws_feat["B2"].font = Font(name=font_name, size=14, bold=True, color="FFFFFF")
ws_feat["B2"].fill = fill_header
ws_feat["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws_feat.row_dimensions[2].height = 35

# Headers
headers_feat = ["Feature ID", "Feature Name", "Description", "Owner", "Status", "Priority", "Completion %", "Comments"]
for col_idx, text in enumerate(headers_feat, start=2):
    cell = ws_feat.cell(row=3, column=col_idx)
    cell.value = text
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_header
ws_feat.row_dimensions[3].height = 25

features_data = [
    ("F-01", "Landing Page", "Landing page explaining NutriBite's clinical and gamified modes", "Pavan Vignesh", "Completed", "Low", "Ready for deployment"),
    ("F-02", "Authentication", "User register, login, JWT and role-based access controls", "Tharun", "Completed", "High", "Verified auth layers"),
    ("F-03", "Parent Dashboard", "Command center for child metrics, sleep, active logs", "Pavan Vignesh", "Completed", "High", "Responsive layout complete"),
    ("F-04", "Doctor Dashboard", "Pediatric dashboard for patients grid, notes and charts", "Pavan Vignesh", "Completed", "High", "Merged details view"),
    ("F-05", "Child Profile Management", "Creation and maintenance of multiple child records", "Tharun", "Completed", "Medium", "Validated with Zod schemas"),
    ("F-06", "Meal Logging", "6 slot meal logging journal with portions and image upload", "Pavan Krishna", "Completed", "High", "Linked to local food db"),
    ("F-07", "Nutrition Analysis", "RDA target comparison and calorie/protein gap tracker", "Pavan Krishna", "Completed", "High", "Deterministic gap analysis"),
    ("F-08", "AI Nutrition Assistant", "Parents conversational RAG guide with citation viewer", "Abhiram", "Completed", "High", "RAG and planning active"),
    ("F-09", "Food Buddy", "Gamified kids chatbot with story prompt filters and XP", "Abhiram", "Completed", "Medium", "Verified kid story filters"),
    ("F-10", "Doctor Access System", "Secure restricted or full permission doctor connection", "Tharun", "Completed", "High", "Validated with handshake logic"),
    ("F-11", "Growth Tracking", "Logs height, weight, BMI risk status over time", "Pavan Krishna", "Completed", "Medium", "Computes BMI percentiles"),
    ("F-12", "RAG Pipeline", "BM25 and FAISS hybrid vector search with reranker", "Dinesh Veera", "Completed", "High", "Guideline textbooks loaded"),
    ("F-13", "Privacy Mode", "Local Ollama offline model router toggle", "Abhiram", "Review", "Medium", "Integrating static fallback"),
    ("F-14", "Deployment", "MongoDB Atlas, Prometheus telemetry and compose setups", "Tharun", "Completed", "High", "Staging containers running")
]

for row_idx, data in enumerate(features_data, start=4):
    fid, fname, desc, owner, status, priority, comments = data
    
    ws_feat.cell(row=row_idx, column=2, value=fid).font = font_regular
    ws_feat.cell(row=row_idx, column=3, value=fname).font = font_bold
    ws_feat.cell(row=row_idx, column=4, value=desc).font = font_regular
    ws_feat.cell(row=row_idx, column=5, value=owner).font = font_regular
    
    # Status
    cell_status = ws_feat.cell(row=row_idx, column=6, value=status)
    cell_status.alignment = Alignment(horizontal="center")
    if status in status_styles:
        cell_status.fill = status_styles[status]["fill"]
        cell_status.font = status_styles[status]["font"]
        
    # Priority
    cell_priority = ws_feat.cell(row=row_idx, column=7, value=priority)
    cell_priority.alignment = Alignment(horizontal="center")
    if priority in priority_styles:
        cell_priority.fill = priority_styles[priority]["fill"]
        cell_priority.font = priority_styles[priority]["font"]
        
    # Completion %
    cell_comp = ws_feat.cell(row=row_idx, column=8)
    cell_comp.value = f'=IF(COUNTIF(Tasks!$C$4:$C$100, C{row_idx})>0, COUNTIFS(Tasks!$C$4:$C$100, C{row_idx}, Tasks!$G$4:$G$100, "Completed")/COUNTIF(Tasks!$C$4:$C$100, C{row_idx}), 0)'
    cell_comp.font = font_bold
    cell_comp.number_format = "0.0%"
    cell_comp.alignment = Alignment(horizontal="right")
    
    ws_feat.cell(row=row_idx, column=9, value=comments).font = font_regular
    
    for c in range(2, 10):
        cell_b = ws_feat.cell(row=row_idx, column=c)
        cell_b.border = border_all
        if row_idx % 2 == 1 and cell_b.fill.fill_type is None:
            cell_b.fill = fill_zebra
            
    ws_feat.row_dimensions[row_idx].height = 22

# Dropdowns for Features
val_status_feat = DataValidation(type="list", formula1='"Open,In Progress,Review,Completed,Blocked"', allow_blank=True)
val_priority_feat = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws_feat.add_data_validation(val_status_feat)
ws_feat.add_data_validation(val_priority_feat)

val_status_feat.add("F4:F17")
val_priority_feat.add("G4:G17")

# ==========================================
# SHEET 3: USER STORIES
# ==========================================
ws_stories = wb.create_sheet(title="User Stories")
ws_stories.views.sheetView[0].showGridLines = True

# Title block
ws_stories.merge_cells("B2:H2")
ws_stories["B2"] = "NutriBite User Stories"
ws_stories["B2"].font = Font(name=font_name, size=14, bold=True, color="FFFFFF")
ws_stories["B2"].fill = fill_header
ws_stories["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws_stories.row_dimensions[2].height = 35

# Headers
headers_stories = ["Story ID", "Feature", "User Story", "Assigned To", "Status", "Reviewer", "Comments"]
for col_idx, text in enumerate(headers_stories, start=2):
    cell = ws_stories.cell(row=3, column=col_idx)
    cell.value = text
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_header
ws_stories.row_dimensions[3].height = 25

stories_data = [
    ("US-01", "Landing Page", "As a prospective user, I want to view a visually engaging landing page, so that I can understand NutriBite's clinical and gamified features.", "Pavan Vignesh", "Completed", "Abhiram", "Visual layout verified"),
    ("US-02", "Authentication", "As a parent or doctor, I want secure registration and login, so that my child's medical and dietary data is protected.", "Tharun", "Completed", "Abhiram", "Auth tests passed"),
    ("US-03", "Parent Dashboard", "As a parent, I want to see a central dashboard with quick logging tools and water intake counters, so that I can easily monitor my child's daily habits.", "Pavan Vignesh", "Completed", "Tharun", "Linked to stats API"),
    ("US-04", "Doctor Dashboard", "As a pediatrician, I want a unified patient list and growth graphs, so that I can review real-time meal diaries and growth histories.", "Pavan Vignesh", "Completed", "Tharun", "Grid layouts complete"),
    ("US-05", "Child Profile Management", "As a parent, I want to create separate child profiles with dietary preferences and custom avatars, so that I can manage each child's health goals individually.", "Tharun", "Completed", "Pavan Krishna", "DB schema tested"),
    ("US-06", "Meal Logging", "As a parent, I want to log breakfast, lunch, dinner, snacks, and upload meal photos, so that I have a complete dietary history for analysis.", "Pavan Krishna", "Completed", "Abhiram", "Form validation tested"),
    ("US-07", "Nutrition Analysis", "As a parent, I want to view auto-calculated macro totals and RDA gaps, so that I can identify dietary deficiencies like iron or calcium.", "Pavan Krishna", "Completed", "Dinesh Veera", "Target comparisons verified"),
    ("US-08", "AI Nutrition Assistant", "As a parent, I want to chat with an AI assistant that references NIN/ICMR guidelines, so that I can get reliable, safe pediatric advice.", "Abhiram", "Completed", "Dinesh Veera", "Citations verified"),
    ("US-09", "Food Buddy", "As a child, I want to chat with a superhero mascot, earn XP, and unlock badges for eating healthy, so that eating well is fun.", "Abhiram", "Completed", "Pavan Vignesh", "Kids mode flow validated"),
    ("US-10", "Doctor Access System", "As a parent, I want to approve or deny access requests from specific doctors, so that I control who sees my child's clinical records.", "Tharun", "Completed", "Abhiram", "Two-way handshake verified"),
    ("US-11", "Growth Tracking", "As a parent, I want to record height and weight logs to calculate BMI, so that I can monitor my child's physical development.", "Pavan Krishna", "Completed", "Tharun", "Growth velocity calculated"),
    ("US-12", "RAG Pipeline", "As a developer, I want hybrid keyword/semantic retrieval, so that chat answers are clinically correct.", "Dinesh Veera", "Completed", "Abhiram", "Retrieval MRR is above 0.85"),
    ("US-13", "Privacy Mode", "As a parent, I want local offline inference routing, so that child PII is not sent to external cloud APIs.", "Abhiram", "Review", "Tharun", "Model router works offline"),
    ("US-14", "Deployment", "As a developer, I want a containerized deployment with metrics, so that operations are stable and observable.", "Tharun", "Completed", "Dinesh Veera", "Telemetry dashboards running")
]

for row_idx, data in enumerate(stories_data, start=4):
    sid, feat, story, assigned, status, reviewer, comments = data
    
    ws_stories.cell(row=row_idx, column=2, value=sid).font = font_regular
    ws_stories.cell(row=row_idx, column=3, value=feat).font = font_bold
    
    cell_story = ws_stories.cell(row=row_idx, column=4, value=story)
    cell_story.font = font_regular
    cell_story.alignment = Alignment(wrap_text=True)
    
    ws_stories.cell(row=row_idx, column=5, value=assigned).font = font_regular
    
    # Status
    cell_status = ws_stories.cell(row=row_idx, column=6, value=status)
    cell_status.alignment = Alignment(horizontal="center")
    if status in status_styles:
        cell_status.fill = status_styles[status]["fill"]
        cell_status.font = status_styles[status]["font"]
        
    ws_stories.cell(row=row_idx, column=7, value=reviewer).font = font_regular
    ws_stories.cell(row=row_idx, column=8, value=comments).font = font_regular
    
    for c in range(2, 9):
        cell_b = ws_stories.cell(row=row_idx, column=c)
        cell_b.border = border_all
        if row_idx % 2 == 1 and cell_b.fill.fill_type is None:
            cell_b.fill = fill_zebra
            
    ws_stories.row_dimensions[row_idx].height = 40

val_status_stories = DataValidation(type="list", formula1='"Open,In Progress,Review,Completed,Blocked"', allow_blank=True)
ws_stories.add_data_validation(val_status_stories)
val_status_stories.add("F4:F17")


# ==========================================
# SHEET 4: TASKS
# ==========================================
ws_tasks = wb.create_sheet(title="Tasks")
ws_tasks.views.sheetView[0].showGridLines = True

# Title block
ws_tasks.merge_cells("B2:I2")
ws_tasks["B2"] = "NutriBite Detailed Task Registry"
ws_tasks["B2"].font = Font(name=font_name, size=14, bold=True, color="FFFFFF")
ws_tasks["B2"].fill = fill_header
ws_tasks["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws_tasks.row_dimensions[2].height = 35

# Headers
headers_tasks = ["Task ID", "Feature", "Task Name", "Assigned To", "Reviewed By", "Status", "Due Date", "Comments"]
for col_idx, text in enumerate(headers_tasks, start=2):
    cell = ws_tasks.cell(row=3, column=col_idx)
    cell.value = text
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_header
ws_tasks.row_dimensions[3].height = 25

tasks_data = [
    # Landing Page
    ("T-001", "Landing Page", "Create Hero Section", "Pavan Vignesh", "Abhiram", "Completed", "2026-04-15", "Responsive layout complete"),
    ("T-002", "Landing Page", "Create About Section", "Pavan Vignesh", "Abhiram", "Completed", "2026-04-16", "Standard copy written"),
    ("T-003", "Landing Page", "Create Features Section", "Pavan Vignesh", "Abhiram", "Completed", "2026-04-17", "Icons added"),
    ("T-004", "Landing Page", "Create Contact Section", "Pavan Vignesh", "Abhiram", "Completed", "2026-04-18", "Feedback form integrated"),
    ("T-005", "Landing Page", "Responsive Design", "Pavan Vignesh", "Abhiram", "Completed", "2026-04-20", "Viewport tested on mobile"),
    # Authentication
    ("T-006", "Authentication", "Login UI", "Pavan Vignesh", "Tharun", "Completed", "2026-04-22", "Auth forms structured"),
    ("T-007", "Authentication", "Registration UI", "Pavan Vignesh", "Tharun", "Completed", "2026-04-23", "Role check dropdown added"),
    ("T-008", "Authentication", "JWT Authentication", "Tharun", "Abhiram", "Completed", "2026-04-25", "Cookie parser configuration"),
    ("T-009", "Authentication", "Role Based Access", "Tharun", "Abhiram", "Completed", "2026-04-27", "Express route guarding"),
    # Parent Dashboard
    ("T-010", "Parent Dashboard", "Child Selection Widget", "Pavan Vignesh", "Tharun", "Completed", "2026-04-29", "Dynamic dropdown list"),
    ("T-011", "Parent Dashboard", "Activity Level Widget", "Pavan Vignesh", "Tharun", "Completed", "2026-04-30", "Active settings tags"),
    ("T-012", "Parent Dashboard", "Caloric Summary UI", "Pavan Vignesh", "Pavan Krishna", "Completed", "2026-05-02", "Circular charts embedded"),
    ("T-013", "Parent Dashboard", "Water Logger Stats Card", "Pavan Krishna", "Tharun", "Completed", "2026-05-04", "Quick +250ml logging"),
    # Doctor Dashboard
    ("T-014", "Doctor Dashboard", "Patients Grid", "Pavan Vignesh", "Tharun", "Completed", "2026-05-06", "Patient list layout complete"),
    ("T-015", "Doctor Dashboard", "Pending Requests Mailbox", "Pavan Vignesh", "Tharun", "Completed", "2026-05-07", "Approval controls"),
    ("T-016", "Doctor Dashboard", "Diagnostic Notes", "Tharun", "Abhiram", "Completed", "2026-05-09", "Notes storage controller"),
    ("T-017", "Doctor Dashboard", "Recharts Growth Curves", "Pavan Vignesh", "Pavan Krishna", "Completed", "2026-05-11", "BMI data plotting"),
    # Child Profile Management
    ("T-018", "Child Profile Management", "Create Profile Form", "Pavan Vignesh", "Tharun", "Completed", "2026-05-13", "React state hooks configured"),
    ("T-019", "Child Profile Management", "Demographics", "Tharun", "Pavan Krishna", "Completed", "2026-05-14", "Mongoose schema defined"),
    ("T-020", "Child Profile Management", "Animal Avatar Selector", "Pavan Vignesh", "Tharun", "Completed", "2026-05-16", "Vibrant graphics selection"),
    ("T-021", "Child Profile Management", "Health Conditions List", "Tharun", "Pavan Krishna", "Completed", "2026-05-17", "Allergy text array"),
    # Meal Logging
    ("T-022", "Meal Logging", "Daily Slots", "Pavan Krishna", "Abhiram", "Completed", "2026-05-19", "6 distinct slot containers"),
    ("T-023", "Meal Logging", "Portion Sliders", "Pavan Vignesh", "Pavan Krishna", "Completed", "2026-05-20", "Vite UI sliders"),
    ("T-024", "Meal Logging", "Photo Upload API", "Tharun", "Abhiram", "Completed", "2026-05-22", "Multer file storage routes"),
    ("T-025", "Meal Logging", "Delete Food Item", "Pavan Krishna", "Tharun", "Completed", "2026-05-23", "Route delete handlers tested"),
    # Nutrition Analysis
    ("T-026", "Nutrition Analysis", "RDA Target Calculation", "Pavan Krishna", "Abhiram", "Completed", "2026-05-25", "WHO reference lookup tables"),
    ("T-027", "Nutrition Analysis", "Gap Detection Algorithm", "Pavan Krishna", "Dinesh Veera", "Completed", "2026-05-26", "Nutrient deficiency checkers"),
    ("T-028", "Nutrition Analysis", "Dynamic Food Recommendations", "Abhiram", "Dinesh Veera", "Completed", "2026-05-28", "Recommender system rules"),
    # AI Nutrition Assistant
    ("T-029", "AI Nutrition Assistant", "Parent Chat Interface", "Pavan Vignesh", "Abhiram", "Completed", "2026-05-30", "Streaming text container"),
    ("T-030", "AI Nutrition Assistant", "Clinical View Toggle", "Pavan Vignesh", "Abhiram", "Completed", "2026-06-01", "Text splitting details"),
    ("T-031", "AI Nutrition Assistant", "Dynamic Query Citations", "Abhiram", "Dinesh Veera", "Completed", "2026-06-02", "Citation source mapping"),
    # Food Buddy
    ("T-032", "Food Buddy", "Avatar Custom Framing", "Pavan Vignesh", "Abhiram", "Completed", "2026-06-04", "Super-hero kids layout theme"),
    ("T-033", "Food Buddy", "Fun Chat Story Prompt Rules", "Abhiram", "Dinesh Veera", "Completed", "2026-06-05", "Interactive kids character logic"),
    ("T-034", "Food Buddy", "XP Tracker", "Pavan Krishna", "Tharun", "Completed", "2026-06-07", "Completed logs reward triggers"),
    ("T-035", "Food Buddy", "Unlocked Badges Drawer", "Pavan Vignesh", "Tharun", "Completed", "2026-06-08", "Badge grid animations"),
    # Doctor Access System
    ("T-036", "Doctor Access System", "Doctor Access Request Link", "Tharun", "Abhiram", "Completed", "2026-06-10", "Email access invitation system"),
    ("T-037", "Doctor Access System", "Parent Access List", "Tharun", "Abhiram", "Completed", "2026-06-11", "Access listing controllers"),
    ("T-038", "Doctor Access System", "Restricted/Full View Approval Logic", "Tharun", "Abhiram", "Completed", "2026-06-12", "Permissions checks middleware"),
    ("T-039", "Doctor Access System", "Revoke Access", "Tharun", "Abhiram", "Completed", "2026-06-13", "Delete access mapping records"),
    # Growth Tracking
    ("T-040", "Growth Tracking", "Height/Weight Logs", "Pavan Krishna", "Tharun", "Completed", "2026-06-14", "Height-weight form fields"),
    ("T-041", "Growth Tracking", "BMI Category Status", "Pavan Krishna", "Tharun", "Completed", "2026-06-15", "Calculates WHO risk statuses"),
    ("T-042", "Growth Tracking", "Longitudinal Growth Database Storage", "Tharun", "Pavan Krishna", "Completed", "2026-06-16", "GrowthRecord schemas indices"),
    # RAG Pipeline
    ("T-043", "RAG Pipeline", "BM25 Keyword Matching", "Dinesh Veera", "Abhiram", "Completed", "2026-05-15", "Okapi BM25 indices created"),
    ("T-044", "RAG Pipeline", "FAISS Dense Search Vector Indexing", "Dinesh Veera", "Abhiram", "Completed", "2026-05-16", "SentenceTransformers indexing"),
    ("T-045", "RAG Pipeline", "Cross-Encoder Reranker", "Dinesh Veera", "Abhiram", "Completed", "2026-05-18", "ms-marco-MiniLM model scoring"),
    ("T-046", "RAG Pipeline", "Textbook Ingestion", "Dinesh Veera", "Abhiram", "Completed", "2026-05-20", "Parsed PDF pages into clean text"),
    # Privacy Mode
    ("T-047", "Privacy Mode", "Offline Toggle", "Pavan Vignesh", "Tharun", "Completed", "2026-06-10", "Dashboard layout toggle button"),
    ("T-048", "Privacy Mode", "Local Ollama Router", "Abhiram", "Dinesh Veera", "Completed", "2026-06-11", "Verifies local tags ping endpoint"),
    ("T-049", "Privacy Mode", "Local Static Fallback Generator", "Abhiram", "Dinesh Veera", "In Progress", "2026-06-20", "Developing backup response template"),
    # Deployment
    ("T-050", "Deployment", "MongoDB Atlas Configuration", "Tharun", "Pavan Krishna", "Completed", "2026-06-14", "Cluster setup and security limits"),
    ("T-051", "Deployment", "NextJS Production Container", "Tharun", "Pavan Krishna", "Completed", "2026-06-15", "Multi-stage build completed"),
    ("T-052", "Deployment", "Prometheus Metrics Scraper", "Dinesh Veera", "Abhiram", "Completed", "2026-06-16", "Instruments AI service metrics"),
    ("T-053", "Deployment", "Docker-Compose Orchestration", "Tharun", "Pavan Krishna", "In Progress", "2026-06-21", "Container bridge configuration")
]

for row_idx, data in enumerate(tasks_data, start=4):
    tid, feat, tname, assigned, reviewer, status, due_date, comments = data
    
    ws_tasks.cell(row=row_idx, column=2, value=tid).font = font_regular
    ws_tasks.cell(row=row_idx, column=3, value=feat).font = font_bold
    ws_tasks.cell(row=row_idx, column=4, value=tname).font = font_regular
    ws_tasks.cell(row=row_idx, column=5, value=assigned).font = font_regular
    ws_tasks.cell(row=row_idx, column=6, value=reviewer).font = font_regular
    
    # Status
    cell_status = ws_tasks.cell(row=row_idx, column=7, value=status)
    cell_status.alignment = Alignment(horizontal="center")
    if status in status_styles:
        cell_status.fill = status_styles[status]["fill"]
        cell_status.font = status_styles[status]["font"]
        
    # Due Date
    cell_date = ws_tasks.cell(row=row_idx, column=8, value=due_date)
    cell_date.alignment = Alignment(horizontal="center")
    cell_date.font = font_regular
    
    ws_tasks.cell(row=row_idx, column=9, value=comments).font = font_regular
    
    for c in range(2, 10):
        cell_b = ws_tasks.cell(row=row_idx, column=c)
        cell_b.border = border_all
        if row_idx % 2 == 1 and cell_b.fill.fill_type is None:
            cell_b.fill = fill_zebra
            
    ws_tasks.row_dimensions[row_idx].height = 20

val_status_tasks = DataValidation(type="list", formula1='"Open,In Progress,Review,Completed,Blocked"', allow_blank=True)
ws_tasks.add_data_validation(val_status_tasks)
val_status_tasks.add(f"G4:G100")


# ==========================================
# SHEET 5: TEAM CONTRIBUTIONS
# ==========================================
ws_team = wb.create_sheet(title="Team Contributions")
ws_team.views.sheetView[0].showGridLines = True

# Title block
ws_team.merge_cells("B2:H2")
ws_team["B2"] = "NutriBite Team Contribution Tracker"
ws_team["B2"].font = Font(name=font_name, size=14, bold=True, color="FFFFFF")
ws_team["B2"].fill = fill_header
ws_team["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws_team.row_dimensions[2].height = 35

# Headers
headers_team = ["Member", "Role", "Features Owned", "Tasks Assigned", "Tasks Completed", "Contribution %", "Comments"]
for col_idx, text in enumerate(headers_team, start=2):
    cell = ws_team.cell(row=3, column=col_idx)
    cell.value = text
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_header
ws_team.row_dimensions[3].height = 25

team_data = [
    ("Abhiram", "AI Engineer & Feature Lead", "LLM Integration, Ollama Router, Story prompts"),
    ("Tharun", "Backend & Database Engineer", "Auth backend, MongoDB database schemas, express server"),
    ("Dinesh Veera", "RAG Engineer", "FAISS indexes, BM25 retriever, Cross-Encoder reranker"),
    ("Pavan Krishna", "Data Engineer & Backend Developer", "Indian foods dataset, sleep log logic, BMI calculations"),
    ("Pavan Vignesh", "Frontend & UI/UX Engineer", "Web UI dashboards, Recharts velocity lines, animations")
]

for idx, data in enumerate(team_data, start=4):
    name, role, comments = data
    
    ws_team.cell(row=idx, column=2, value=name).font = font_bold
    ws_team.cell(row=idx, column=3, value=role).font = font_regular
    
    # Features Owned
    cell_f = ws_team.cell(row=idx, column=4, value=f'=COUNTIF(Features!$E$4:$E$17, B{idx})')
    cell_f.font = font_regular
    cell_f.alignment = Alignment(horizontal="center")
    
    # Tasks Assigned
    cell_t = ws_team.cell(row=idx, column=5, value=f'=COUNTIF(Tasks!$E$4:$E$100, B{idx})')
    cell_t.font = font_regular
    cell_t.alignment = Alignment(horizontal="center")
    
    # Tasks Completed
    cell_c = ws_team.cell(row=idx, column=6, value=f'=COUNTIFS(Tasks!$E$4:$E$100, B{idx}, Tasks!$G$4:$G$100, "Completed")')
    cell_c.font = font_regular
    cell_c.alignment = Alignment(horizontal="center")
    
    # Contribution %
    cell_p = ws_team.cell(row=idx, column=7, value=f'=IF(SUM($F$4:$F$8)>0, F{idx}/SUM($F$4:$F$8), 0)')
    cell_p.font = font_bold
    cell_p.number_format = "0.0%"
    cell_p.alignment = Alignment(horizontal="right")
    
    ws_team.cell(row=idx, column=8, value=comments).font = font_regular
    
    for c in range(2, 9):
        cell_b = ws_team.cell(row=idx, column=c)
        cell_b.border = border_all
        if idx % 2 == 1 and cell_b.fill.fill_type is None:
            cell_b.fill = fill_zebra
            
    ws_team.row_dimensions[idx].height = 22

# Total Row in Team Contributions
ws_team.cell(row=9, column=2, value="Total").font = font_bold
ws_team.cell(row=9, column=2).border = border_double_bottom
ws_team.cell(row=9, column=3, value="").border = border_double_bottom

ws_team.cell(row=9, column=4, value="=SUM(D4:D8)").font = font_bold
ws_team.cell(row=9, column=4).border = border_double_bottom
ws_team.cell(row=9, column=4).alignment = Alignment(horizontal="center")

ws_team.cell(row=9, column=5, value="=SUM(E4:E8)").font = font_bold
ws_team.cell(row=9, column=5).border = border_double_bottom
ws_team.cell(row=9, column=5).alignment = Alignment(horizontal="center")

ws_team.cell(row=9, column=6, value="=SUM(F4:F8)").font = font_bold
ws_team.cell(row=9, column=6).border = border_double_bottom
ws_team.cell(row=9, column=6).alignment = Alignment(horizontal="center")

ws_team.cell(row=9, column=7, value="=SUM(G4:G8)").font = font_bold
ws_team.cell(row=9, column=7).border = border_double_bottom
ws_team.cell(row=9, column=7).alignment = Alignment(horizontal="right")
ws_team.cell(row=9, column=7).number_format = "0.0%"

ws_team.cell(row=9, column=8, value="").border = border_double_bottom
ws_team.row_dimensions[9].height = 22


# ==========================================
# COLUMN WIDTHS CONFIGURATION
# ==========================================
col_widths = {
    "Dashboard": {1: 2, 2: 22, 3: 22, 4: 22, 5: 22, 6: 22, 7: 22},
    "Features": {1: 2, 2: 12, 3: 28, 4: 55, 5: 20, 6: 15, 7: 12, 8: 15, 9: 30},
    "User Stories": {1: 2, 2: 12, 3: 25, 4: 65, 5: 20, 6: 15, 7: 20, 8: 30},
    "Tasks": {1: 2, 2: 10, 3: 25, 4: 45, 5: 20, 6: 20, 7: 15, 8: 14, 9: 30},
    "Team Contributions": {1: 2, 2: 20, 3: 32, 4: 16, 5: 16, 6: 16, 7: 18, 8: 50}
}

for sheet_name, widths in col_widths.items():
    ws = wb[sheet_name]
    for col_idx, width in widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

# Save file
wb.save("NutriBite_Project_Tracker.xlsx")
print("Successfully generated NutriBite_Project_Tracker.xlsx!")
